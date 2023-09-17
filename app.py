from flask import Flask, redirect, render_template, request, Response, redirect
from flask_sqlalchemy import SQLAlchemy
import cv2
from datetime import datetime as dt
import face_recognition
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# name of flask app as app
app = Flask(__name__)

# configure it with a database named as register which will store the information of user
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///register.db'
db = SQLAlchemy(app)

# Here we are using two year excel file of 2022 and 2023  
year = ['2022', '2023']
month = [['January', 31], ['February', 28], ['March', 31], ['April', 30], ['May', 31],
         ['June', 30], ['July', 31], ['August', 31], ['September', 30], ['October', 31],
          ['November', 30], ['December', 31]]


# Creation of the table of the attendance of date and month only to be executed once to create the table
# It can also be done manually we just create number of workbooks as 1 per year with each sheet storing a month
# for example here we have create for the 2022 and 2023 workbook 
# for x in range(2):
#     wb = Workbook('Attendance'+year[x]+'.xlsx')
#     for i in range(12):
#         ws = wb.create_sheet(month[i][0])
#         ws.append(['EmailID']+ [str(y+1) for y in range(month[i][1])])
#     wb.save('Attendance'+year[x]+'.xlsx')


# the database object will store the name, email , password, face encodings and the registration time info
class User(db.Model):
    email = db.Column(db.String(200), primary_key=True)
    password = db.Column(db.String(100), nullable=False)
    user_name = db.Column(db.String(100), nullable=False)
    encoding = db.Column(db.String(), nullable=False)
    registration_date = db.Column(db.DateTime, default=dt.utcnow)

    def __repr__(self) -> str:
        # The object will return the email and password if we use the print function on it
        return "{}-{}".format(self.email,self.password)

# It sets the bunch of global declared values as zero it is like asynchronous resetting the values
def global_reset()-> None:
        global face_encoding
        face_encoding = []
        global user_email
        user_email = ""
        global name_user
        name_user = ""
        global user_password
        user_password = ""
        global match
        match = ""
        global user
        user = None
        global count
        count = 0

# There is a text file with name AdminPassword.txt which stores the password and can password is retrieved from it
def get_admin_password()-> str:
    with open('AdminPassword.txt') as f:
        admin_password = f.read()
        f.close()
        return admin_password

# It add a new row of registration of a new enrolled user in all the defined attendance workbook and month sheets using
# there email id
def add_registration(user_email:str) -> None:
    for x in range(len(year)):
        wb = load_workbook('Attendance'+year[x]+'.xlsx')
        for i in range(12):
            ws = wb[month[i][0]]
            ws.append([user_email])
        wb.save('Attendance'+year[x]+'.xlsx')
    return

# delete a registration in the excel file of the registered user
def delete_registration(user_email:str) -> None:
    wb = load_workbook('Attendance2022.xlsx')
    ws = wb[month[1][0]]
    row= 2
    while ws['A'+str(row)].value != None:
        if ws['A'+str(row)].value == user_email:
            break
        row+=1
    wb.save('Attendance2022.xlsx')
    for x in range(len(year)):
        wb = load_workbook('Attendance'+year[x]+'.xlsx')
        for i in range(12):
            ws = wb[month[i][0]]
            ws.delete_rows(row)
        wb.save('Attendance'+year[x]+'.xlsx')
    return None

# Search for the cell address where to put the attendance and returns it as a string
def get_cell_address(user_email)-> str:
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    time = list_dt[1].split('.')[0]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[month_choice-1][0]]
    col_id = get_column_letter(date+1)
    row= 2
    val = None
    while ws['A'+str(row)].value != None:
        if ws['A'+str(row)].value == user_email:
            val = ws[col_id+str(row)].value
            break
        row += 1
    wb.save('Attendance'+str(file_year)+'.xlsx')
    return (col_id + str(row), val)

# According to the current date in the respective workbook it sets the attendance as P
def set_attendance(user_email)-> None:
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[month_choice-1][0]]
    ws[get_cell_address(user_email)[0]].value = "P"
    wb.save('Attendance'+str(file_year)+'.xlsx')
    return 
    
# While registering t the camera captures the image and return it to the image box as .jpg at every moment
def  generate_frame ():
    # global variables are created so that this can be accessible to other function 
    global face_encoding
    global camera
    camera = cv2.VideoCapture(0)
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # Resizing is done to reduce the amount of processing 
            frame2 = cv2.resize(frame, (0,0), None, 0.25, 0.25)
            faceLoc = face_recognition.face_locations(frame2)
            face_encoding = face_recognition.face_encodings(frame2)
            if len(faceLoc)!=0:
                faceLoc = faceLoc[0]
                faceLoc2 = [int(x*4) for x in faceLoc]
                # Create the rectangle on the face if face is recognized
                frame = cv2.rectangle(frame, (faceLoc2[3], faceLoc2[0]), (faceLoc2[1], faceLoc2[2]), (255, 0, 0), 1)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
        # Constantly provide the image to the image box of the html page
        yield(b'--frame\r\n'b'Content-Type : image/jpeg\r\n\r\n'+frame+b'\r\n')

# It is to recognize the current user face while running the program and return it to the image box as .jpg at every moment
def  generate_frame_compare():
    global count
    count = 0
    global camera
    camera = cv2.VideoCapture(0)
    registered_encoding = [float(x) for x in user.encoding.split()]
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            frame2 = cv2.resize(frame, (0,0), None, 0.25, 0.25)
            faceLoc = face_recognition.face_locations(frame2)
            face_encoding = face_recognition.face_encodings(frame2,faceLoc)
            # Compare only when there is some face on the screen
            if len(faceLoc)!=0:
                faceLoc = faceLoc[0]
                faceLoc2 = [int(x*4) for x in faceLoc]
                # Compare the face data of the camera image wth the previously stored data in the database about the face
                result = face_recognition.compare_faces([registered_encoding],face_encoding[0],tolerance=0.4)
                if True in result:
                    count +=1
                    # if the user is match his name appears on the image 
                    frame = cv2.putText(frame,user.user_name, (faceLoc2[3]+6, faceLoc2[2]-6),  cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1 )
                    if count >= 8:
                        # when the face is matched about 8 times then the attendance is accepted here and stop image
                        set_attendance(user.email)
                        camera.release()
                        frame = cv2.putText(frame,"Face Matched\n Click Done", (faceLoc2[3]+6, faceLoc2[2]-6),  cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1 )
                        ret, buffer = cv2.imencode('.jpg', frame)
                        frame = buffer.tobytes()
                        return b'--frame\r\n'b'Content-Type : image/jpeg\r\n\r\n'+frame+b'\r\n'
                else:
                    frame = cv2.putText(frame,"User Do Not Match", (faceLoc2[3]+6, faceLoc2[2]-6),  cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1 )
                frame = cv2.rectangle(frame, (faceLoc2[3], faceLoc2[0]), (faceLoc2[1], faceLoc2[2]), (255, 0, 0), 1)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
        # Constantly provide the image to the image box of the html page
        yield(b'--frame\r\n'b'Content-Type : image/jpeg\r\n\r\n'+frame+b'\r\n')

# The home page is routed to the index page
@app.route('/')
def index():
    # The main home poge of the web app is called
    try:
        camera.release()
        if name_user == "" and match==1:
            global_reset()
            # Opens the main page with a text message when the registration is complete
            return render_template('index.html', tasks= 0)
        else:
            global_reset()
            return render_template('index.html',tasks=1)
    except: 
        pass
    global_reset()
    return render_template('index.html', tasks=1)

# Get the registration details of the user 
@app.route('/register', methods=['POST', 'GET'])
def register():
    error = 0
    if request.method == 'POST':
        global user_email
        user_email = request.form['email_id']
        global name_user
        name_user = request.form['user_id']
        global user_password
        user_password = request.form['password']
        global match
        match = request.form['password_match']

        if len(user_email.strip())==0 or len(name_user.strip())==0 or len(user_password.strip())==0 or len(match.strip()) == 0:
            error = 1
            # Error because the fields were empty so ask the user to enter the information again and put a message
            return render_template('register.html', tasks=error)
        elif user_password != match:
            error = 2
             # Error because the password did not match so ask the user to enter the information again and put a message
            return render_template('register.html', tasks=error)
        else:
            try:
                 # Error because the email id already exist so ask the user to enter the information again and put a message
                User.query.get_or_404(user_email)
                error = 3
                return render_template('register.html', tasks=error)
            except:
                # If the criteria of registration is met then it movers to face registration
                return redirect('/registercam')
    else:
        return render_template ('register.html', tasks=error)

# Loads a webpage where camera registration takes place and is used to 
@app.route('/registercam', methods=['POST', 'GET'])
def registercam():
    if request.method =="POST":
        # Loads the page where the face of the user is being registered
        if len(face_encoding)==0:
            # if the face is not recognized to come out with a message
            return render_template('registerCam.html', tasks=0)
        else:
            # converted the user face encoding into string to store
            encoding_list = face_encoding[0].tolist()
            encoded_string = " ".join([str(X) for X in encoding_list])
            # create an object to store in the database
            user_register = User(email=user_email, password=user_password, user_name=name_user, encoding=encoded_string)
            # Add the registered data of the user in the database
            db.session.add(user_register)
            db.session.commit()
            add_registration(user_email)
            global_reset()
            # Here match is used as a flag to show that the registration is being done
            global match
            match=1
            return redirect('/')
        
    else:
        return render_template('registerCam.html',tasks=1)

# It renders the frames into the image box in the html page
@app.route('/video')
def video():
    # get the camera input in the browser 
    return Response(generate_frame(), mimetype='multipart/x-mixed-replace; boundary=frame')


# It is the login page of the user with their email and password
@app.route('/login', methods = ['POST', 'GET'])
def login():
    if request.method == 'POST':
        login_id = request.form['login_id']
        login_password = request.form['login_password']
        try:
            global user
            user = User.query.get_or_404(login_id)
            if user.password == login_password:
                # if the password is match and moves to the user page
                return redirect('/login/user')
            else:
                # Reloads the same page with a message of incorrect password
                return render_template('login.html', tasks=0)
        except:
            # Reloads the same page with the message of invalid email ID
            return render_template('login.html', tasks=1)
        
    else:
        return render_template('login.html', tasks=2)

# The user page page to interact and perform action like view and put attendance is loaded
@app.route('/login/user')
def login_user():
    return render_template('user.html', tasks = [user.user_name, user.email])

# When the attendance is done This page appears representing attendance of today is done
@app.route('/completed')
def completed():
    try:
        camera.release()
    except:
        pass
    global_reset()
    return render_template('completed.html')

# IT loads the face recognition page of verfying the attendance for that day
@app.route('/attendance', methods=['POST', 'GET'])
def attendance():
    if request.method =='POST':
        if count >= 8:
            # When the attendance is done and button is pressed then it takes to completed page
            return redirect('/completed')
        else:
            # If the Done button is pressed before the face recognition 
            return render_template('put_attendance.html', tasks=0)
    else:  
        if user is None: 
            return redirect('/login/user')
        elif get_cell_address(user.email)[1]=='P':
            # If the attendance is already done directly shows the completed page
            return redirect('/completed')
        # Loads the user page
        return render_template('put_attendance.html')

# The image is rendered in the attendance image frame using it 
@app.route('/video_attend')
def video_attend():
    if user is None:
        return redirect('/login/user')
    # get the camera input in the browser 
    return Response(generate_frame_compare(), mimetype='multipart/x-mixed-replace; boundary=frame')

# The User details to see which months attendance is to be seen the information is being loaded
@app.route('/view')
def view():
    if user is None:
        return redirect('/login')
    return render_template('month.html', tasks = [user.user_name, user.email, str(user.registration_date).split()[0]])

# The month attendance is shwon in form of table using the data from the excel sheet
@app.route('/month_view/<int:id>')
def month_view(id):
    # It searches for the row of the user email id in the excel sheet
    list_dt = str(dt.now()).split()
    file_year, month_choice, date = [int(X) for X in list_dt[0].split('-')]
    wb = load_workbook('Attendance'+str(file_year)+'.xlsx')
    ws = wb[month[id-1][0]]
    row= 2
    val = []
    while ws['A'+str(row)].value != None:
        if ws['A'+str(row)].value == user.email:
            break
        row+=1
    # For the selected month it get the value of attendance in the array
    for col in range(2, int(month[id-1][1])+2):
        if ws[get_column_letter(col)+str(row)].value =='P':
            val.append("Present")
        else:
            val.append("Not Available")
    wb.save('Attendance'+str(file_year)+'.xlsx')
    tasks = [user.user_name, user.email, str(user.registration_date).split()[0], month[id-1], val]
    return render_template('view_attendance.html', tasks=tasks)

# This loads the admin password page so that he can lagin
@app.route('/admin_login', methods=['POST', 'GET'])
def admin_login():
    if request.method == 'POST':
        pass_ad = request.form['admin_password']
        if pass_ad == get_admin_password():
            # GOes to the admin control page if the password is matched 
            return redirect('/admin')
        else:
            # if the password is wrong it reloads the admin login page with the message
            return render_template('admin_login.html', tasks=1)
    else:
        # Loads the password input page
        return render_template('admin_login.html', task=0)

# The admin control page is loaded when it is called
@app.route('/admin')
def admin():
    return render_template('admin_control.html')

# The page is loaded with name and email of registered persons and can view one's monthly attendance
@app.route('/admin_view', methods=['POST', 'GET'])
def admin_view():
    if request.method == 'POST':
        user_email = request.form['useremail']
        try:
            # If the admin enters an correct email id of the user he goes to see his attendance
            x = User.query.get_or_404(user_email)
            global user
            user = x
            return redirect('/view')
        except:
            # if the entered email ID don't exist then reload the page with an message
            tasks = User.query.order_by(User.registration_date).all()
            return render_template('admin_view.html',data=0,tasks=tasks)
    else:
        tasks = User.query.order_by(User.registration_date).all()
        return render_template('admin_view.html',data=1, tasks=tasks)

# The page is loaded with the name and email and the delete using the email id of that user
@app.route('/delete', methods=['POST', 'GET'])
def delete():
    if request.method == 'POST':
        user_email = request.form['useremail']
        try:
            # Deletes the user from the database as well as his information from the excel is deleted
            x = User.query.get_or_404(user_email)
            # Function to delete user from the excel sheet is called 
            delete_registration(x.email)
            # Function to delete the user registration from the database
            db.session.delete(x)
            db.session.commit()
            return redirect('/delete')
        except:
            # If invalid email is entered the same page is reloaded with the message
            tasks = User.query.order_by(User.registration_date).all()
            return render_template('admin_delete.html', data=0,tasks=tasks)
    else:
        tasks = User.query.order_by(User.registration_date).all()
        return render_template('admin_delete.html', data=1, tasks=tasks)

# Run the app using the app.py file
if __name__ == '__main__':
    app.run()